import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

REPORTS_DIR = Path(__file__).parent.parent / "shipping_reports"
INDEX_PATH = REPORTS_DIR / "index.json"

_reports_index: list[dict] = []
_counter: int = 0


def _load_index():
    global _reports_index, _counter
    if INDEX_PATH.exists():
        with open(INDEX_PATH, "r", encoding="utf-8") as f:
            _reports_index = json.load(f)
        _counter = len(_reports_index)


def _save_index():
    INDEX_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(INDEX_PATH, "w", encoding="utf-8") as f:
        json.dump(_reports_index, f, ensure_ascii=False, indent=2)


def _type_label(report_type: str) -> str:
    return {"daily": "日报", "weekly": "周报", "monthly": "月报"}.get(report_type, report_type)


def generate_report(
    report_date: str,
    report_type: str,
    start_date: str,
    end_date: str,
    customer_name: str | None = None,
    product_name: str | None = None,
) -> dict:
    """内部查询发货数据，生成 Excel 报表并存入 shipping_reports 目录"""
    from services import shipping_service

    # 空字符串视为无筛选
    customer_name = customer_name.strip() if customer_name else None
    product_name = product_name.strip() if product_name else None

    _load_index()

    # 0. 去重：同一 report_date + report_type + 筛选条件 已存在则直接返回
    for meta in _reports_index:
        if (meta["report_date"] == report_date
                and meta["report_type"] == report_type
                and meta.get("customer_name") == customer_name
                and meta.get("product_name") == product_name):
            return {
                "success": True,
                "file_name": meta["file_name"],
                "file_path": meta["file_path"],
                "storage_id": meta["storage_id"],
                "download_url": meta["download_url"],
                "message": "该报表已存在，直接返回",
            }

    # 1. 内部查询发货数据
    shipping_data = shipping_service.query_shipping_list(
        start_date, end_date, customer_name, product_name
    )
    records = shipping_data["records"]
    summary = shipping_data["summary"]

    if not records:
        return {"success": False, "message": "该时间段内无发货记录，无法生成报表"}

    # 2. 生成报表文件
    global _counter
    _counter += 1

    label = _type_label(report_type)
    suffix = ""
    if customer_name:
        suffix += f"_{customer_name}"
    if product_name:
        suffix += f"_{product_name}"
    file_name = f"发货{label}_{report_date}{suffix}.xlsx"
    year_month = report_date[:7].replace("-", "/").split("/")
    sub_dir = REPORTS_DIR / year_month[0] / year_month[1]
    sub_dir.mkdir(parents=True, exist_ok=True)
    file_path = sub_dir / file_name
    storage_id = f"RPT_{report_date.replace('-', '')}_{_counter:03d}"

    # 生成 Excel
    wb = Workbook()

    # Sheet1: 发货明细
    ws1 = wb.active
    ws1.title = "发货明细"
    headers = ["序号", "日期", "发货单号", "客户", "产品", "规格", "方量", "重量", "金额"]
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for i, rec in enumerate(records, 2):
        row_data = [
            i - 1,
            rec.get("shipping_date", ""),
            rec.get("shipping_order_no", ""),
            rec.get("customer_name", ""),
            rec.get("product_name", ""),
            rec.get("spec_model", ""),
            rec.get("cubic_volume", 0),
            rec.get("quantity", 0),
            rec.get("total_price_with_tax", 0),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=i, column=col, value=val)
            cell.border = thin_border

    # Sheet2: 汇总统计
    ws2 = wb.create_sheet("汇总统计")

    # 总汇总
    ws2.cell(row=1, column=1, value=f"报表日期：{report_date}").font = Font(bold=True, size=13)
    ws2.cell(row=2, column=1, value=f"数据范围：{start_date} ~ {end_date}")
    ws2.cell(row=3, column=1, value=f"总订单数：{summary['total_records']}")
    ws2.cell(row=4, column=1, value=f"总方量：{summary['total_cubic_volume']} m³")
    ws2.cell(row=5, column=1, value=f"总重量：{summary['total_quantity']} 吨")
    ws2.cell(row=6, column=1, value=f"总金额：¥{summary['total_price_with_tax']:,.2f}")

    # 按客户汇总
    from collections import defaultdict
    customer_stats = defaultdict(lambda: {"orders": 0, "cubic_volume": 0.0, "quantity": 0.0, "amount": 0.0})
    for rec in records:
        c = customer_stats[rec["customer_name"]]
        c["orders"] += 1
        c["cubic_volume"] += rec.get("cubic_volume", 0)
        c["quantity"] += rec.get("quantity", 0)
        c["amount"] += rec.get("total_price_with_tax", 0)

    row = 8
    ws2.cell(row=row, column=1, value="按客户汇总").font = Font(bold=True, size=12)
    row += 1
    cust_headers = ["客户名称", "订单数", "方量(m³)", "重量(吨)", "金额(元)"]
    for col, h in enumerate(cust_headers, 1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for cust_name in sorted(customer_stats.keys()):
        row += 1
        s = customer_stats[cust_name]
        cust_row = [cust_name, s["orders"], round(s["cubic_volume"], 4), round(s["quantity"], 4), round(s["amount"], 2)]
        for col, val in enumerate(cust_row, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.border = thin_border

    # 按产品汇总
    product_stats = defaultdict(lambda: {"orders": 0, "cubic_volume": 0.0, "quantity": 0.0, "amount": 0.0})
    for rec in records:
        p = product_stats[rec["product_name"]]
        p["orders"] += 1
        p["cubic_volume"] += rec.get("cubic_volume", 0)
        p["quantity"] += rec.get("quantity", 0)
        p["amount"] += rec.get("total_price_with_tax", 0)

    row += 2
    ws2.cell(row=row, column=1, value="按产品汇总").font = Font(bold=True, size=12)
    row += 1
    prod_headers = ["产品名称", "订单数", "方量(m³)", "重量(吨)", "金额(元)"]
    for col, h in enumerate(prod_headers, 1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for prod_name in sorted(product_stats.keys()):
        row += 1
        s = product_stats[prod_name]
        prod_row = [prod_name, s["orders"], round(s["cubic_volume"], 4), round(s["quantity"], 4), round(s["amount"], 2)]
        for col, val in enumerate(prod_row, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.border = thin_border

    # 调整列宽
    ws2.column_dimensions["A"].width = 35

    wb.save(str(file_path))

    # 更新索引
    meta = {
        "storage_id": storage_id,
        "file_name": file_name,
        "report_date": report_date,
        "report_type": report_type,
        "customer_name": customer_name,
        "product_name": product_name,
        "created_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        "file_path": str(file_path.relative_to(REPORTS_DIR.parent)),
        "download_url": f"/api/excel/download/{storage_id}",
    }
    _reports_index.append(meta)
    _save_index()

    return {
        "success": True,
        "file_name": file_name,
        "file_path": str(file_path.relative_to(REPORTS_DIR.parent)),
        "storage_id": storage_id,
        "download_url": f"/api/excel/download/{storage_id}",
    }


def query_reports(start_date: str, end_date: str, report_type: str | None = None) -> list[dict]:
    """按日期范围查询报表元数据"""
    _load_index()
    results = []
    for meta in _reports_index:
        if start_date <= meta["report_date"] <= end_date:
            if report_type and meta["report_type"] != report_type:
                continue
            results.append({
                "storage_id": meta["storage_id"],
                "file_name": meta["file_name"],
                "report_date": meta["report_date"],
                "report_type": meta["report_type"],
                "created_at": meta["created_at"],
                "download_url": meta["download_url"],
            })
    return results


def get_file_path(storage_id: str) -> str | None:
    """根据 storage_id 获取报表文件的物理路径"""
    _load_index()
    for meta in _reports_index:
        if meta["storage_id"] == storage_id:
            full_path = REPORTS_DIR.parent / meta["file_path"]
            if full_path.exists():
                return str(full_path)
    return None
